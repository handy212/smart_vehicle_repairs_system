from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Max, Min, Avg, F, Case, When, Value, CharField
from django.db import models
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from datetime import datetime, timedelta

from .models import (
    Part, Supplier, PurchaseOrder, PurchaseOrderItem, 
    PartCategory, InventoryTransaction
)
from apps.accounts.models import User


@login_required
def part_list_view(request):
    """
    Parts catalog with advanced search and filtering
    """
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager', 'technician']:
        messages.error(request, 'You do not have permission to view parts.')
        return redirect('dashboard:home')
    
    # Get filters
    search = request.GET.get('search', '')
    category_id = request.GET.get('category')
    supplier_id = request.GET.get('supplier')
    status = request.GET.get('status')  # 'active' or 'inactive'
    stock_level = request.GET.get('stock_level')
    sort_by = request.GET.get('sort', '-created_at')
    
    # Base queryset
    parts = Part.objects.select_related(
        'category', 'preferred_supplier'
    ).prefetch_related('suppliers').annotate(
        available_stock=F('quantity_in_stock') - F('quantity_reserved'),
        stock_status=Case(
            When(quantity_in_stock__lte=F('minimum_stock'), then=Value('low')),
            When(quantity_in_stock__lte=F('reorder_point'), then=Value('reorder')),
            default=Value('normal'),
            output_field=CharField()
        )
    )
    
    # Apply filters
    if search:
        parts = parts.filter(
            Q(name__icontains=search) | 
            Q(part_number__icontains=search) |
            Q(manufacturer__icontains=search) |
            Q(description__icontains=search)
        )
    
    if category_id:
        parts = parts.filter(category_id=category_id)
        
    if supplier_id:
        parts = parts.filter(suppliers__id=supplier_id)
        
    # Default to showing only active parts unless status filter is explicitly set
    if status == 'inactive':
        parts = parts.filter(is_active=False)
    elif status == 'all':
        # Show both active and inactive
        pass
    else:
        # Default to active only (includes when status is 'active' or not provided)
        parts = parts.filter(is_active=True)
    
    if stock_level:
        if stock_level == 'low':
            parts = parts.filter(quantity_in_stock__lte=F('minimum_stock'))
        elif stock_level == 'reorder':
            parts = parts.filter(quantity_in_stock__lte=F('reorder_point'))
        elif stock_level == 'out':
            parts = parts.filter(quantity_in_stock=0)
    
    # Sorting
    if sort_by in ['-created_at', 'name', '-name', 'part_number', '-part_number', 
                   'quantity_in_stock', '-quantity_in_stock', 'cost_price', '-cost_price']:
        parts = parts.order_by(sort_by)
    else:
        parts = parts.order_by('-created_at')
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format in ['csv', 'pdf']:
        return export_parts(parts, export_format)
    
    # Pagination - 20 items per page, with option to load more (40)
    per_page = int(request.GET.get('per_page', 20))
    if per_page not in [20, 40]:
        per_page = 20
    
    paginator = Paginator(parts, per_page)
    page = request.GET.get('page')
    parts = paginator.get_page(page)
    
    # Get filter options
    categories = PartCategory.objects.order_by('name')
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')
    
    # Statistics
    total_parts = Part.objects.count()
    low_stock_count = Part.objects.filter(quantity_in_stock__lte=F('minimum_stock')).count()
    out_of_stock_count = Part.objects.filter(quantity_in_stock=0).count()
    
    context = {
        'parts': parts,
        'categories': categories,
        'suppliers': suppliers,
        'search': search,
        'category_id': int(category_id) if category_id else None,
        'supplier_id': int(supplier_id) if supplier_id else None,
        'status': status,
        'stock_level': stock_level,
        'sort_by': sort_by,
        'total_parts': total_parts,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'can_edit': request.user.role in ['admin', 'manager', 'parts_manager'],
        'per_page': per_page,
    }
    
    return render(request, 'inventory/part_list.html', context)


def export_parts(queryset, format_type):
    """
    Export parts to CSV or PDF
    """
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="parts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Part Number', 'Name', 'Category', 'Supplier', 'In Stock', 
            'Unit', 'Price', 'Status'
        ])
        
        for part in queryset:
            writer.writerow([
                part.part_number,
                part.name,
                part.category.name if part.category else '',
                part.preferred_supplier.name if part.preferred_supplier else '',
                part.quantity_in_stock,
                part.get_unit_display(),
                float(part.selling_price) if part.selling_price else '',
                'Active' if part.is_active else 'Inactive'
            ])
        
        return response
    
    elif format_type == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="parts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title = Paragraph(f"<b>Parts Inventory Export</b><br/>{datetime.now().strftime('%B %d, %Y')}", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.25*inch))
        
        # Table data
        data = [['Part #', 'Name', 'Category', 'Supplier', 'Stock', 'Unit', 'Price', 'Status']]
        
        for part in queryset[:100]:  # Limit to 100 for PDF
            data.append([
                str(part.part_number)[:15],
                str(part.name)[:30],
                str(part.category.name)[:20] if part.category else '-',
                str(part.preferred_supplier.name)[:20] if part.preferred_supplier else '-',
                str(part.quantity_in_stock),
                str(part.get_unit_display())[:8],
                f"${float(part.selling_price):.2f}" if part.selling_price else '-',
                'Active' if part.is_active else 'Inactive'
            ])
        
        # Create table
        table = Table(data, colWidths=[0.8*inch, 2*inch, 1.2*inch, 1.5*inch, 0.6*inch, 0.6*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return response


@login_required
def part_import_view(request):
    """
    Import parts from CSV or Excel file
    """
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to import parts.')
        return redirect('inventory:part_list')
    
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'No file uploaded.')
            return redirect('inventory:part_import')
        
        file = request.FILES['file']
        file_ext = file.name.split('.')[-1].lower()
        
        if file_ext not in ['csv', 'xlsx', 'xls']:
            messages.error(request, 'Invalid file format. Please upload CSV or Excel file.')
            return redirect('inventory:part_import')
        
        try:
            imported_count = 0
            updated_count = 0
            error_rows = []
            
            if file_ext == 'csv':
                # Handle CSV import
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        part_number = row.get('part_number', '').strip()
                        name = row.get('name', '').strip()
                        
                        if not part_number or not name:
                            error_rows.append(f"Row {row_num}: Missing part_number or name")
                            continue
                        
                        # Get or create category
                        category = None
                        if row.get('category'):
                            category, _ = PartCategory.objects.get_or_create(
                                name=row['category'].strip(),
                                defaults={'created_by': request.user}
                            )
                        
                        # Get or create supplier
                        supplier = None
                        if row.get('supplier'):
                            supplier, _ = Supplier.objects.get_or_create(
                                name=row['supplier'].strip(),
                                defaults={'created_by': request.user}
                            )
                        
                        # Create or update part
                        part, created = Part.objects.update_or_create(
                            part_number=part_number,
                            defaults={
                                'name': name,
                                'description': row.get('description', ''),
                                'category': category,
                                'preferred_supplier': supplier,
                                'manufacturer': row.get('manufacturer', ''),
                                'manufacturer_part_number': row.get('manufacturer_part_number', ''),
                                'quantity_in_stock': int(row.get('quantity_in_stock', 0)),
                                'minimum_stock': int(row.get('minimum_stock', 0)),
                                'reorder_point': int(row.get('reorder_point', 0)),
                                'reorder_quantity': int(row.get('reorder_quantity', 0)),
                                'cost_price': float(row.get('cost_price', 0)) if row.get('cost_price') else None,
                                'selling_price': float(row.get('selling_price', 0)) if row.get('selling_price') else None,
                                'bin_location': row.get('bin_location', ''),
                                'shelf': row.get('shelf', ''),
                                'unit': row.get('unit', 'unit'),
                                'is_active': row.get('is_active', 'true').lower() in ['true', '1', 'yes'],
                                'created_by': request.user
                            }
                        )
                        
                        if created:
                            imported_count += 1
                        else:
                            updated_count += 1
                            
                    except Exception as e:
                        error_rows.append(f"Row {row_num}: {str(e)}")
            
            else:
                # Handle Excel import
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                
                # Get headers from first row
                headers = [cell.value for cell in ws[1]]
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        data = dict(zip(headers, row))
                        
                        part_number = str(data.get('part_number', '')).strip()
                        name = str(data.get('name', '')).strip()
                        
                        if not part_number or not name:
                            error_rows.append(f"Row {row_num}: Missing part_number or name")
                            continue
                        
                        # Get or create category
                        category = None
                        if data.get('category'):
                            category, _ = PartCategory.objects.get_or_create(
                                name=str(data['category']).strip(),
                                defaults={'created_by': request.user}
                            )
                        
                        # Get or create supplier
                        supplier = None
                        if data.get('supplier'):
                            supplier, _ = Supplier.objects.get_or_create(
                                name=str(data['supplier']).strip(),
                                defaults={'created_by': request.user}
                            )
                        
                        # Create or update part
                        part, created = Part.objects.update_or_create(
                            part_number=part_number,
                            defaults={
                                'name': name,
                                'description': str(data.get('description', '')),
                                'category': category,
                                'preferred_supplier': supplier,
                                'manufacturer': str(data.get('manufacturer', '')),
                                'manufacturer_part_number': str(data.get('manufacturer_part_number', '')),
                                'quantity_in_stock': int(data.get('quantity_in_stock', 0)) if data.get('quantity_in_stock') else 0,
                                'minimum_stock': int(data.get('minimum_stock', 0)) if data.get('minimum_stock') else 0,
                                'reorder_point': int(data.get('reorder_point', 0)) if data.get('reorder_point') else 0,
                                'reorder_quantity': int(data.get('reorder_quantity', 0)) if data.get('reorder_quantity') else 0,
                                'cost_price': float(data.get('cost_price', 0)) if data.get('cost_price') else None,
                                'selling_price': float(data.get('selling_price', 0)) if data.get('selling_price') else None,
                                'bin_location': str(data.get('bin_location', '')),
                                'shelf': str(data.get('shelf', '')),
                                'unit': str(data.get('unit', 'unit')),
                                'is_active': str(data.get('is_active', 'true')).lower() in ['true', '1', 'yes'],
                                'created_by': request.user
                            }
                        )
                        
                        if created:
                            imported_count += 1
                        else:
                            updated_count += 1
                            
                    except Exception as e:
                        error_rows.append(f"Row {row_num}: {str(e)}")
            
            # Show results
            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} new parts.')
            if updated_count > 0:
                messages.info(request, f'Updated {updated_count} existing parts.')
            if error_rows:
                messages.warning(request, f'{len(error_rows)} rows had errors. Check the error log.')
                # Store errors in session for display
                request.session['import_errors'] = error_rows[:50]  # Limit to 50 errors
            
            return redirect('inventory:part_list')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('inventory:part_import')
    
    # GET request - show import form
    context = {
        'can_edit': True,
        'import_errors': request.session.pop('import_errors', [])
    }
    return render(request, 'inventory/part_import.html', context)


@login_required
def part_import_template_view(request):
    """
    Download CSV template for importing parts
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="parts_import_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'part_number', 'name', 'description', 'category', 'supplier', 
        'manufacturer', 'manufacturer_part_number', 'quantity_in_stock', 
        'minimum_stock', 'reorder_point', 'reorder_quantity', 'cost_price', 
        'selling_price', 'bin_location', 'shelf', 'unit', 'is_active'
    ])
    
    # Add sample rows
    writer.writerow([
        'PART-001', 'Sample Oil Filter', 'High quality oil filter', 'Filters', 
        'ABC Supplier', 'Bosch', 'BOF-123', '50', '10', '15', '25', 
        '5.50', '12.99', 'A-12', 'Shelf 3', 'unit', 'true'
    ])
    writer.writerow([
        'PART-002', 'Brake Pad Set', 'Front brake pads', 'Brakes', 
        'XYZ Parts', 'Brembo', 'BP-456', '25', '5', '10', '15', 
        '35.00', '89.99', 'B-05', 'Shelf 1', 'box', 'true'
    ])
    
    return response


@login_required
def part_detail_view(request, pk):
    """
    Part detail view with stock history and transactions
    """
    part = get_object_or_404(
        Part.objects.select_related('category', 'preferred_supplier')
        .prefetch_related('suppliers', 'transactions__created_by'),
        pk=pk
    )
    
    # Get recent transactions
    recent_transactions = part.transactions.select_related('created_by').order_by('-created_at')[:20]
    
    # Calculate statistics
    total_purchased = part.transactions.filter(transaction_type='purchase').aggregate(
        total=Sum('quantity')
    )['total'] or 0
    
    total_used = part.transactions.filter(transaction_type='sale').aggregate(
        total=Sum('quantity')
    )['total'] or 0
    
    # Get reorder suggestions
    needs_reorder = part.quantity_in_stock <= part.reorder_point
    
    # Calculate markup and inventory value
    markup_amount = None
    markup_percentage = None
    inventory_value = None
    
    if part.cost_price and part.selling_price:
        markup_amount = float(part.selling_price) - float(part.cost_price)
        if part.cost_price > 0:
            markup_percentage = (markup_amount / float(part.cost_price)) * 100
    
    if part.cost_price and part.quantity_in_stock:
        inventory_value = float(part.cost_price) * part.quantity_in_stock
    
    context = {
        'part': part,
        'recent_transactions': recent_transactions,
        'total_purchased': total_purchased,
        'total_used': abs(total_used),  # Make positive for display
        'needs_reorder': needs_reorder,
        'available_stock': part.quantity_in_stock - part.quantity_reserved,
        'can_edit': request.user.role in ['admin', 'manager', 'parts_manager'],
        'markup_amount': markup_amount,
        'markup_percentage': markup_percentage,
        'inventory_value': inventory_value,
    }
    
    return render(request, 'inventory/part_detail.html', context)


@login_required
def part_create_view(request):
    """
    Create new part
    """
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to create parts.')
        return redirect('inventory:part_list')
    
    categories = PartCategory.objects.order_by('name')
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')
    
    if request.method == 'POST':
        try:
            # Create part
            part = Part.objects.create(
                name=request.POST.get('name'),
                part_number=request.POST.get('part_number'),
                description=request.POST.get('description', ''),
                category_id=request.POST.get('category') or None,
                manufacturer=request.POST.get('manufacturer', ''),
                manufacturer_part_number=request.POST.get('manufacturer_part_number', ''),
                cost_price=float(request.POST.get('cost_price', 0)),
                selling_price=float(request.POST.get('selling_price', 0)),
                quantity_in_stock=int(request.POST.get('quantity_in_stock', 0)),
                reorder_point=int(request.POST.get('reorder_point', 10)),
                reorder_quantity=int(request.POST.get('reorder_quantity', 20)),
                minimum_stock=int(request.POST.get('minimum_stock', 5)),
                maximum_stock=int(request.POST.get('maximum_stock')) if request.POST.get('maximum_stock') else None,
                unit=request.POST.get('unit', 'piece'),
                bin_location=request.POST.get('bin_location', ''),
                shelf=request.POST.get('shelf', ''),
                weight=float(request.POST.get('weight')) if request.POST.get('weight') else None,
                warranty_months=int(request.POST.get('warranty_months')) if request.POST.get('warranty_months') else None,
                is_active=request.POST.get('is_active') == 'on',
                is_taxable=request.POST.get('is_taxable') == 'on',
                is_core=request.POST.get('is_core') == 'on',
                preferred_supplier_id=request.POST.get('preferred_supplier') or None,
                created_by=request.user
            )
            
            # Handle image upload
            if 'image' in request.FILES:
                part.image = request.FILES['image']
                part.save()
            
            # Add suppliers
            supplier_ids = request.POST.getlist('suppliers')
            if supplier_ids:
                part.suppliers.set(supplier_ids)
            
            # Create initial stock transaction
            if part.quantity_in_stock > 0:
                InventoryTransaction.objects.create(
                    part=part,
                    transaction_type='adjustment',
                    quantity=part.quantity_in_stock,
                    balance_after=part.quantity_in_stock,
                    unit_cost=part.cost_price,
                    total_cost=part.cost_price * part.quantity_in_stock,
                    reason='Initial stock',
                    notes='Initial stock entry',
                    created_by=request.user
                )
            
            messages.success(request, f'Part "{part.name}" created successfully!')
            return redirect('inventory:part_detail', pk=part.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating part: {str(e)}')
    
    context = {
        'categories': categories,
        'suppliers': suppliers,
        'UNIT_CHOICES': Part.UNIT_CHOICES,
    }
    
    return render(request, 'inventory/part_create_simple.html', context)


@login_required
def part_edit_view(request, pk):
    """
    Edit existing part
    """
    part = get_object_or_404(Part, pk=pk)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to edit parts.')
        return redirect('inventory:part_detail', pk=part.pk)
    
    categories = PartCategory.objects.order_by('name')
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')
    
    if request.method == 'POST':
        try:
            # Track stock changes
            old_stock = part.quantity_in_stock
            new_stock = int(request.POST.get('quantity_in_stock', 0))
            
            # Update part
            part.name = request.POST.get('name')
            part.part_number = request.POST.get('part_number')
            part.description = request.POST.get('description', '')
            part.category_id = request.POST.get('category') or None
            part.manufacturer = request.POST.get('manufacturer', '')
            part.manufacturer_part_number = request.POST.get('manufacturer_part_number', '')
            part.cost_price = float(request.POST.get('cost_price', 0))
            part.selling_price = float(request.POST.get('selling_price', 0))
            part.quantity_in_stock = new_stock
            part.reorder_point = int(request.POST.get('reorder_point', 10))
            part.reorder_quantity = int(request.POST.get('reorder_quantity', 20))
            part.minimum_stock = int(request.POST.get('minimum_stock', 5))
            part.maximum_stock = int(request.POST.get('maximum_stock')) if request.POST.get('maximum_stock') else None
            part.unit = request.POST.get('unit', 'piece')
            part.bin_location = request.POST.get('bin_location', '')
            part.shelf = request.POST.get('shelf', '')
            part.weight = float(request.POST.get('weight')) if request.POST.get('weight') else None
            part.warranty_months = int(request.POST.get('warranty_months')) if request.POST.get('warranty_months') else None
            part.is_active = request.POST.get('is_active') == 'on'
            part.is_taxable = request.POST.get('is_taxable') == 'on'
            part.is_core = request.POST.get('is_core') == 'on'
            part.preferred_supplier_id = request.POST.get('preferred_supplier') or None
            
            # Handle image upload
            if 'image' in request.FILES:
                part.image = request.FILES['image']
            
            part.save()
            
            # Update suppliers
            supplier_ids = request.POST.getlist('suppliers')
            part.suppliers.set(supplier_ids)
            
            # Create stock adjustment transaction if quantity changed
            if new_stock != old_stock:
                quantity_change = new_stock - old_stock
                InventoryTransaction.objects.create(
                    part=part,
                    transaction_type='adjustment',
                    quantity=quantity_change,
                    balance_after=new_stock,
                    unit_cost=part.cost_price,
                    total_cost=part.cost_price * abs(quantity_change),
                    reason='Manual adjustment',
                    notes=f'Stock adjusted from {old_stock} to {new_stock}',
                    created_by=request.user
                )
            
            messages.success(request, f'Part "{part.name}" updated successfully!')
            return redirect('inventory:part_detail', pk=part.pk)
            
        except Exception as e:
            messages.error(request, f'Error updating part: {str(e)}')
    
    context = {
        'part': part,
        'categories': categories,
        'suppliers': suppliers,
        'UNIT_CHOICES': Part.UNIT_CHOICES,
    }
    
    return render(request, 'inventory/part_edit.html', context)


@login_required
def supplier_list_view(request):
    """
    Supplier list with search and filtering
    """
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to view suppliers.')
        return redirect('dashboard:home')
    
    search = request.GET.get('search', '')
    supplier_type = request.GET.get('type')
    is_active = request.GET.get('active')
    sort_by = request.GET.get('sort', 'name')
    
    # Base queryset
    suppliers = Supplier.objects.annotate(
        parts_count=Count('parts'),
        purchase_orders_count=Count('purchase_orders')
    )
    
    # Apply filters
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(supplier_code__icontains=search) |
            Q(contact_person__icontains=search) |
            Q(email__icontains=search)
        )
    
    if supplier_type:
        suppliers = suppliers.filter(supplier_type=supplier_type)
    
    if is_active:
        suppliers = suppliers.filter(is_active=is_active == 'true')
    
    # Sorting
    if sort_by in ['name', '-name', 'supplier_code', '-supplier_code', 
                   'supplier_type', '-supplier_type', 'created_at', '-created_at']:
        suppliers = suppliers.order_by(sort_by)
    else:
        suppliers = suppliers.order_by('name')
    
    # Pagination
    paginator = Paginator(suppliers, 20)
    page = request.GET.get('page')
    suppliers = paginator.get_page(page)
    
    context = {
        'suppliers': suppliers,
        'search': search,
        'supplier_type': supplier_type,
        'is_active': is_active,
        'sort_by': sort_by,
        'SUPPLIER_TYPE_CHOICES': Supplier.SUPPLIER_TYPE_CHOICES,
        'can_edit': request.user.role in ['admin', 'manager', 'parts_manager'],
    }
    
    return render(request, 'inventory/supplier_list.html', context)


@login_required
def purchase_order_list_view(request):
    """
    Purchase order list with filtering
    """
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to view purchase orders.')
        return redirect('dashboard:home')
    
    search = request.GET.get('search', '')
    status = request.GET.get('status')
    supplier_id = request.GET.get('supplier')
    sort_by = request.GET.get('sort', '-created_at')
    
    # Base queryset
    purchase_orders = PurchaseOrder.objects.select_related(
        'supplier', 'created_by'
    ).annotate(
        items_count=Count('items'),
        total_quantity=Sum('items__quantity'),
        received_quantity=Sum('items__quantity_received')
    )
    
    # Apply filters
    if search:
        purchase_orders = purchase_orders.filter(
            Q(po_number__icontains=search) |
            Q(supplier__name__icontains=search) |
            Q(notes__icontains=search)
        )
    
    if status:
        purchase_orders = purchase_orders.filter(status=status)
    
    if supplier_id:
        purchase_orders = purchase_orders.filter(supplier_id=supplier_id)
    
    # Sorting
    if sort_by in ['-created_at', 'po_number', '-po_number', 'order_date', 
                   '-order_date', 'total', '-total', 'status', '-status']:
        purchase_orders = purchase_orders.order_by(sort_by)
    else:
        purchase_orders = purchase_orders.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(purchase_orders, 20)
    page = request.GET.get('page')
    purchase_orders = paginator.get_page(page)
    
    # Get filter options
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')
    
    context = {
        'purchase_orders': purchase_orders,
        'suppliers': suppliers,
        'search': search,
        'status': status,
        'supplier_id': int(supplier_id) if supplier_id else None,
        'sort_by': sort_by,
        'STATUS_CHOICES': PurchaseOrder.STATUS_CHOICES,
        'can_edit': request.user.role in ['admin', 'manager', 'parts_manager'],
    }
    
    return render(request, 'inventory/purchase_order_list.html', context)


@login_required
def purchase_order_detail_view(request, pk):
    """
    Purchase order detail view
    """
    purchase_order = get_object_or_404(
        PurchaseOrder.objects.select_related('supplier', 'created_by')
        .prefetch_related('items__part'),
        pk=pk
    )
    
    context = {
        'purchase_order': purchase_order,
        'can_edit': request.user.role in ['admin', 'manager', 'parts_manager'],
    }
    
    return render(request, 'inventory/purchase_order_detail.html', context)


@login_required
def purchase_order_create_view(request):
    """
    Create new purchase order
    """
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to create purchase orders.')
        return redirect('inventory:purchase_order_list')
    
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')
    parts = Part.objects.filter(is_active=True).order_by('name')
    
    if request.method == 'POST':
        try:
            # Create purchase order
            po = PurchaseOrder.objects.create(
                supplier_id=request.POST.get('supplier'),
                order_date=request.POST.get('order_date'),
                expected_delivery_date=request.POST.get('expected_delivery_date') or None,
                shipping_cost=float(request.POST.get('shipping_cost', 0)),
                tax_amount=float(request.POST.get('tax_amount', 0)),
                notes=request.POST.get('notes', ''),
                internal_notes=request.POST.get('internal_notes', ''),
                created_by=request.user
            )
            
            # Add items
            part_ids = request.POST.getlist('part_id[]')
            quantities = request.POST.getlist('quantity[]')
            unit_costs = request.POST.getlist('unit_cost[]')
            
            for i, part_id in enumerate(part_ids):
                if part_id and quantities[i] and unit_costs[i]:
                    PurchaseOrderItem.objects.create(
                        purchase_order=po,
                        part_id=part_id,
                        quantity=int(quantities[i]),
                        unit_cost=float(unit_costs[i])
                    )
                    
                    # Update part quantity_on_order
                    part = Part.objects.get(id=part_id)
                    part.quantity_on_order += int(quantities[i])
                    part.save()
            
            # Calculate totals
            po.calculate_totals()
            
            messages.success(request, f'Purchase order {po.po_number} created successfully!')
            return redirect('inventory:purchase_order_detail', pk=po.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating purchase order: {str(e)}')
    
    context = {
        'suppliers': suppliers,
        'parts': parts,
    }
    
    return render(request, 'inventory/purchase_order_create.html', context)


@login_required
@require_http_methods(["POST"])
def adjust_stock(request, pk):
    """
    Adjust part stock level
    """
    part = get_object_or_404(Part, pk=pk)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        adjustment_type = request.POST.get('adjustment_type', 'adjustment')
        quantity_change = int(request.POST.get('quantity_change', 0))
        reason = request.POST.get('reason', '')
        notes = request.POST.get('notes', '')
        
        if quantity_change == 0:
            return JsonResponse({'success': False, 'error': 'Quantity change cannot be zero'})
        
        # Calculate new balance
        old_balance = part.quantity_in_stock
        new_balance = max(0, old_balance + quantity_change)
        actual_change = new_balance - old_balance
        
        # Update part stock
        part.quantity_in_stock = new_balance
        part.save()
        
        # Create transaction
        InventoryTransaction.objects.create(
            part=part,
            transaction_type=adjustment_type,
            quantity=actual_change,
            balance_after=new_balance,
            unit_cost=part.cost_price,
            total_cost=part.cost_price * abs(actual_change),
            reason=reason,
            notes=notes,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Stock adjusted successfully. New balance: {new_balance}',
            'new_balance': new_balance
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def supplier_import_view(request):
    """
    Supplier import functionality (placeholder)
    """
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to import suppliers.')
        return redirect('inventory:supplier_list')
    
    if request.method == 'POST':
        messages.info(request, 'Supplier import functionality coming soon.')
        return redirect('inventory:supplier_list')
    
    return render(request, 'inventory/supplier_import.html')


@login_required
def inventory_dashboard_view(request):
    """
    Inventory dashboard with key metrics
    """
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to view inventory dashboard.')
        return redirect('dashboard:home')
    
    # Key metrics
    total_parts = Part.objects.count()
    total_value = Part.objects.aggregate(
        total=Sum(F('quantity_in_stock') * F('cost_price'))
    )['total'] or 0
    
    low_stock_parts = Part.objects.filter(quantity_in_stock__lte=F('minimum_stock'))
    low_stock_count = low_stock_parts.count()
    
    out_of_stock_count = Part.objects.filter(quantity_in_stock=0).count()
    
    reorder_needed = Part.objects.filter(quantity_in_stock__lte=F('reorder_point'))
    reorder_count = reorder_needed.count()
    
    # Recent transactions
    recent_transactions = InventoryTransaction.objects.select_related(
        'part', 'created_by'
    ).order_by('-created_at')[:10]
    
    # Pending purchase orders
    pending_pos = PurchaseOrder.objects.filter(
        status__in=['draft', 'submitted', 'confirmed']
    ).select_related('supplier').order_by('-created_at')[:5]
    
    context = {
        'total_parts': total_parts,
        'total_value': total_value,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'reorder_count': reorder_count,
        'low_stock_parts': low_stock_parts[:10],
        'reorder_needed': reorder_needed[:10],
        'recent_transactions': recent_transactions,
        'pending_pos': pending_pos,
    }
    
    return render(request, 'inventory/dashboard.html', context)


@login_required
@require_http_methods(["POST"])
def part_delete_view(request, pk):
    """Delete a part"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to delete parts.')
        return redirect('inventory:part_list')
    
    part = get_object_or_404(Part, pk=pk)
    force_delete = request.POST.get('force_delete') == 'true'
    
    # If not force delete, check for safety
    if not force_delete:
        has_issues = False
        issues = []
        
        # Check if part has transactions
        if part.transactions.exists():
            has_issues = True
            issues.append(f'{part.transactions.count()} transaction(s)')
        
        # Check if part is in purchase orders
        if hasattr(part, 'purchaseorderitem_set') and part.purchaseorderitem_set.exists():
            has_issues = True
            issues.append(f'{part.purchaseorderitem_set.count()} purchase order(s)')
        
        # Check if part has stock
        if part.quantity_in_stock > 0:
            has_issues = True
            issues.append(f'{part.quantity_in_stock} units in stock')
        
        if has_issues:
            # Store in session for confirmation page
            request.session['delete_part_id'] = pk
            request.session['delete_part_issues'] = issues
            messages.warning(
                request,
                f'"{part.name}" has: {", ".join(issues)}. This will permanently delete all associated data. '
                f'Are you sure you want to continue?'
            )
            return redirect('inventory:part_delete_confirm', pk=pk)
    
    # Proceed with deletion
    part_name = part.name
    
    # Delete all related transactions first
    if part.transactions.exists():
        part.transactions.all().delete()
    
    part.delete()
    
    if force_delete:
        messages.success(request, f'Part "{part_name}" and all related data have been permanently deleted.')
    else:
        messages.success(request, f'Part "{part_name}" has been deleted successfully.')
    
    return redirect('inventory:part_list')


@login_required
def part_delete_confirm_view(request, pk):
    """Confirmation page for deleting a part with data"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to delete parts.')
        return redirect('inventory:part_list')
    
    part = get_object_or_404(Part, pk=pk)
    
    # Get issues from session
    issues = request.session.get('delete_part_issues', [])
    
    context = {
        'part': part,
        'issues': issues,
        'transaction_count': part.transactions.count(),
        'can_edit': True,
    }
    
    return render(request, 'inventory/part_delete_confirm.html', context)


@login_required
def supplier_create_view(request):
    """Create a new supplier"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to create suppliers.')
        return redirect('inventory:supplier_list')
    
    if request.method == 'POST':
        try:
            supplier = Supplier.objects.create(
                name=request.POST.get('name'),
                supplier_code=request.POST.get('supplier_code'),
                supplier_type=request.POST.get('supplier_type', 'distributor'),
                contact_person=request.POST.get('contact_person', ''),
                email=request.POST.get('email', ''),
                phone=request.POST.get('phone', ''),
                address_line1=request.POST.get('address_line1', ''),
                city=request.POST.get('city', ''),
                state=request.POST.get('state', ''),
                postal_code=request.POST.get('postal_code', ''),
                country=request.POST.get('country', 'USA'),
                payment_terms=request.POST.get('payment_terms', ''),
                is_active=request.POST.get('is_active') == 'on',
                is_preferred=request.POST.get('is_preferred') == 'on',
                notes=request.POST.get('notes', ''),
                created_by=request.user
            )
            messages.success(request, f'Supplier "{supplier.name}" created successfully.')
            return redirect('inventory:supplier_detail', pk=supplier.pk)
        except Exception as e:
            messages.error(request, f'Error creating supplier: {str(e)}')
    
    supplier_types = Supplier.SUPPLIER_TYPE_CHOICES
    return render(request, 'inventory/supplier_create.html', {'supplier_types': supplier_types})


@login_required
def supplier_detail_view(request, pk):
    """View supplier details"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager', 'technician']:
        messages.error(request, 'You do not have permission to view suppliers.')
        return redirect('dashboard:home')
    
    supplier = get_object_or_404(
        Supplier.objects.prefetch_related('parts', 'purchase_orders'),
        pk=pk
    )
    
    # Get parts from this supplier
    parts = supplier.parts.all()[:20]
    
    # Get recent purchase orders
    purchase_orders = supplier.purchase_orders.order_by('-created_at')[:10]
    
    context = {
        'supplier': supplier,
        'parts': parts,
        'purchase_orders': purchase_orders,
    }
    
    return render(request, 'inventory/supplier_detail.html', context)


@login_required
def supplier_edit_view(request, pk):
    """Edit a supplier"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to edit suppliers.')
        return redirect('inventory:supplier_list')
    
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        try:
            supplier.name = request.POST.get('name')
            supplier.supplier_code = request.POST.get('supplier_code')
            supplier.supplier_type = request.POST.get('supplier_type', 'distributor')
            supplier.contact_person = request.POST.get('contact_person', '')
            supplier.email = request.POST.get('email', '')
            supplier.phone = request.POST.get('phone', '')
            supplier.address_line1 = request.POST.get('address_line1', '')
            supplier.city = request.POST.get('city', '')
            supplier.state = request.POST.get('state', '')
            supplier.postal_code = request.POST.get('postal_code', '')
            supplier.country = request.POST.get('country', 'USA')
            supplier.payment_terms = request.POST.get('payment_terms', '')
            supplier.is_active = request.POST.get('is_active') == 'on'
            supplier.is_preferred = request.POST.get('is_preferred') == 'on'
            supplier.notes = request.POST.get('notes', '')
            supplier.save()
            
            messages.success(request, f'Supplier "{supplier.name}" updated successfully.')
            return redirect('inventory:supplier_detail', pk=supplier.pk)
        except Exception as e:
            messages.error(request, f'Error updating supplier: {str(e)}')
    
    supplier_types = Supplier.SUPPLIER_TYPE_CHOICES
    context = {
        'supplier': supplier,
        'supplier_types': supplier_types,
    }
    return render(request, 'inventory/supplier_edit.html', context)


@login_required
@require_http_methods(["POST"])
def supplier_delete_view(request, pk):
    """Delete a supplier"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to delete suppliers.')
        return redirect('inventory:supplier_list')
    
    supplier = get_object_or_404(Supplier, pk=pk)
    
    # Check if supplier has parts or purchase orders
    if supplier.parts.exists():
        messages.error(request, f'Cannot delete {supplier.name}. It has associated parts.')
        return redirect('inventory:supplier_detail', pk=pk)
    
    if supplier.purchase_orders.exists():
        messages.error(request, f'Cannot delete {supplier.name}. It has purchase orders.')
        return redirect('inventory:supplier_detail', pk=pk)
    
    supplier_name = supplier.name
    supplier.delete()
    messages.success(request, f'Supplier "{supplier_name}" has been deleted successfully.')
    return redirect('inventory:supplier_list')


@login_required
def purchase_order_edit_view(request, pk):
    """Edit a purchase order"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to edit purchase orders.')
        return redirect('inventory:purchase_order_list')
    
    po = get_object_or_404(PurchaseOrder.objects.prefetch_related('items__part'), pk=pk)
    
    # Can only edit draft purchase orders
    if po.status != 'draft':
        messages.error(request, 'Can only edit draft purchase orders.')
        return redirect('inventory:purchase_order_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            po.supplier_id = request.POST.get('supplier_id')
            po.expected_delivery_date = request.POST.get('expected_delivery_date') or None
            po.notes = request.POST.get('notes', '')
            po.save()
            
            messages.success(request, f'Purchase Order {po.po_number} updated successfully.')
            return redirect('inventory:purchase_order_detail', pk=po.pk)
        except Exception as e:
            messages.error(request, f'Error updating purchase order: {str(e)}')
    
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')
    context = {
        'po': po,
        'suppliers': suppliers,
    }
    return render(request, 'inventory/purchase_order_edit.html', context)


@login_required
@require_http_methods(["POST"])
def purchase_order_delete_view(request, pk):
    """Delete a purchase order"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to delete purchase orders.')
        return redirect('inventory:purchase_order_list')
    
    po = get_object_or_404(PurchaseOrder, pk=pk)
    
    # Can only delete draft purchase orders
    if po.status != 'draft':
        messages.error(request, 'Can only delete draft purchase orders.')
        return redirect('inventory:purchase_order_detail', pk=pk)
    
    po_number = po.po_number
    po.delete()
    messages.success(request, f'Purchase Order {po_number} has been deleted successfully.')
    return redirect('inventory:purchase_order_list')


@login_required
def category_list_view(request):
    """List all part categories"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager', 'technician']:
        messages.error(request, 'You do not have permission to view categories.')
        return redirect('dashboard:home')
    
    categories = PartCategory.objects.annotate(
        parts_count=Count('parts')
    ).order_by('name')
    
    context = {
        'categories': categories,
    }
    return render(request, 'inventory/category_list.html', context)


@login_required
def category_create_view(request):
    """Create a new category"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to create categories.')
        return redirect('inventory:category_list')
    
    if request.method == 'POST':
        try:
            category = PartCategory.objects.create(
                name=request.POST.get('name'),
                description=request.POST.get('description', ''),
                parent_id=request.POST.get('parent_id') or None,
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, f'Category "{category.name}" created successfully.')
            return redirect('inventory:category_list')
        except Exception as e:
            messages.error(request, f'Error creating category: {str(e)}')
    
    categories = PartCategory.objects.filter(is_active=True).order_by('name')
    context = {
        'categories': categories,
    }
    return render(request, 'inventory/category_create.html', context)


@login_required
def category_edit_view(request, pk):
    """Edit a category"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to edit categories.')
        return redirect('inventory:category_list')
    
    category = get_object_or_404(PartCategory, pk=pk)
    
    if request.method == 'POST':
        try:
            category.name = request.POST.get('name')
            category.description = request.POST.get('description', '')
            category.parent_id = request.POST.get('parent_id') or None
            category.is_active = request.POST.get('is_active') == 'on'
            category.save()
            
            messages.success(request, f'Category "{category.name}" updated successfully.')
            return redirect('inventory:category_list')
        except Exception as e:
            messages.error(request, f'Error updating category: {str(e)}')
    
    categories = PartCategory.objects.filter(is_active=True).exclude(pk=pk).order_by('name')
    context = {
        'category': category,
        'categories': categories,
    }
    return render(request, 'inventory/category_edit.html', context)


@login_required
@require_http_methods(["POST"])
def category_delete_view(request, pk):
    """Delete a category"""
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'parts_manager']:
        messages.error(request, 'You do not have permission to delete categories.')
        return redirect('inventory:category_list')
    
    category = get_object_or_404(PartCategory, pk=pk)
    
    # Check if category has parts
    if category.parts.exists():
        messages.error(request, f'Cannot delete {category.name}. It has associated parts.')
        return redirect('inventory:category_list')
    
    # Check if category has subcategories
    if category.subcategories.exists():
        messages.error(request, f'Cannot delete {category.name}. It has subcategories.')
        return redirect('inventory:category_list')
    
    category_name = category.name
    category.delete()
    messages.success(request, f'Category "{category_name}" has been deleted successfully.')
    return redirect('inventory:category_list')