from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, F, Max
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from decimal import Decimal
import json
import csv
from io import BytesIO
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from apps.branches.utils import resolve_branch, filter_queryset_for_user_branches

from .models import (
    WorkOrder, ServiceTask, WorkOrderPart, 
    TechnicianTimeLog, WorkOrderNote, WorkOrderPhoto
)
from apps.diagnosis.models import Diagnosis, RepairRecommendation
from apps.customers.models import Customer
from apps.vehicles.models import Vehicle
from apps.accounts.models import User
from apps.billing.models import Estimate, EstimateLineItem


WORKORDER_EXPORT_HEADERS = [
    'Date In', 'Work Order Number', 'Registration Number', 'Customer Code',
    'Customer Name', 'Customer Balance', 'Job Status', 'Phone Number',
    'Mileage In', 'Job Description', 'Contact Name', 'Received By', 'Needs By',
    'Payment Method', 'Created By', 'Work Hours', 'Labor Amount',
    'Parts Amount', 'Sublet Amount', 'Misc Amount', 'Paid Amount',
    'Tax Amount', 'Date Out', 'Last Note Date', 'Last Note', 'Status Date',
    'VAT Rate', 'Other Tax Rate', 'User Code', 'Changed Date',
    'Recommendations', 'Regular Service', 'Region Code', 'Region', 'Follow Up',
    'Feedback', 'Account To Bill', 'Service Category', 'Invoice Status',
    'Job Type', 'Approval/Diagnosis Comment', 'Repair By', 'Job Location',
    'Insurance Provider', 'Next Action By', 'Next Action On', 'Job Notes',
    'Parts Required', 'Parts Progress', 'Close Date', 'Approval Requested Date',
    'Approved Date', 'Job Hours', 'Invoice Comment', 'Customer Score',
    'Approval Status', 'Subcontract Amount', 'Test Drive By',
    'Quality/Diagnosis Comment', 'Return Reference Number', 'Invoice Date',
    'Mileage In/Out', 'Make', 'Model',
]


def _export_date(value):
    if not value:
        return ''
    if hasattr(value, 'date') and not isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return value.strftime('%Y-%m-%d')


def _export_datetime(value):
    if not value:
        return ''
    return value.strftime('%Y-%m-%d %H:%M:%S')


def _export_decimal(value):
    if value is None or value == '':
        return ''
    try:
        return f"{Decimal(str(value)):.2f}"
    except Exception:
        return str(value)


def _export_bool(value):
    return 'Y' if value else 'N'


def _full_name(user):
    return user.get_full_name() if user else ''


def _customer_name(customer):
    if not customer:
        return ''
    if customer.company_name:
        return customer.company_name
    user = getattr(customer, 'user', None)
    return _full_name(user)


def _contact_name(customer):
    if not customer:
        return ''
    return customer.contact_person_name or _full_name(getattr(customer, 'user', None))


def _customer_phone(customer):
    user = getattr(customer, 'user', None) if customer else None
    return (
        getattr(user, 'phone', '')
        or getattr(customer, 'company_phone', '')
        or getattr(customer, 'alternative_phone', '')
        or ''
    )


def _payment_method(customer):
    if not customer or not customer.default_payment_method:
        return ''
    return customer.get_default_payment_method_display()


def _latest_note(workorder):
    note = workorder.notes.order_by('-created_at').first()
    return note.note if note else ''


def _latest_note_datetime(workorder):
    note = workorder.notes.order_by('-created_at').first()
    return note.created_at if note else None


def _parts_required(workorder):
    return '; '.join(
        filter(None, workorder.parts.order_by('created_at').values_list('part_name', flat=True))
    )


def _parts_progress(workorder):
    statuses = list(workorder.parts.order_by('created_at').values_list('status', flat=True))
    if not statuses:
        return ''

    counts = {}
    for status in statuses:
        counts[status] = counts.get(status, 0) + 1
    return '; '.join(f'{status}:{count}' for status, count in counts.items())


def _recommendations_summary(workorder):
    try:
        diagnosis = workorder.diagnosis
    except Diagnosis.DoesNotExist:
        diagnosis = None
    if not diagnosis:
        return ''
    descriptions = diagnosis.repair_recommendations.order_by('created_at').values_list('description', flat=True)
    return '; '.join(filter(None, descriptions))


def _approval_summary(workorder):
    if workorder.approved_by_customer:
        method = workorder.get_approval_method_display() if workorder.approval_method else 'Approved'
        return method
    if workorder.requires_approval:
        return 'Pending'
    return 'Not Required'


def _resolved_workorder_labor_hours(workorder):
    actual_hours = getattr(workorder, 'actual_labor_hours', None)
    if actual_hours and Decimal(str(actual_hours)) > 0:
        return actual_hours
    totals = workorder.time_logs.aggregate(total=Sum('duration_hours'))
    return totals['total'] or Decimal('0')


def _resolved_workorder_labor_amount(workorder):
    actual_cost = getattr(workorder, 'actual_labor_cost', None)
    if actual_cost and Decimal(str(actual_cost)) > 0:
        return actual_cost
    totals = workorder.time_logs.aggregate(total=Sum('labor_cost'))
    return totals['total'] or Decimal('0')


def _resolved_workorder_parts_amount(workorder):
    actual_parts = getattr(workorder, 'actual_parts_cost', None)
    if actual_parts and Decimal(str(actual_parts)) > 0:
        return actual_parts
    totals = workorder.parts.aggregate(total=Sum('selling_price'))
    return totals['total'] or Decimal('0')


def _primary_invoice_for_export(workorder):
    try:
        from apps.billing.work_order_invoices import get_primary_invoice
    except ImportError:
        return None
    return get_primary_invoice(workorder)


def _tax_rates_for_export():
    try:
        from apps.billing.tax_service import TaxService
    except ImportError:
        return '', ''

    config = TaxService.get_config()
    tax1 = _export_decimal(config.vat_rate)
    tax2 = _export_decimal(config.nhil_rate + config.getfund_rate)
    return tax1, tax2


def _workorder_export_row(workorder, tax1_rate='', tax2_rate=''):
    invoice = _primary_invoice_for_export(workorder)
    customer = getattr(workorder, 'customer', None)
    vehicle = getattr(workorder, 'vehicle', None)
    branch = getattr(workorder, 'branch', None)
    try:
        triage = workorder.triage_form
    except WorkOrder.triage_form.RelatedObjectDoesNotExist:
        triage = None
    latest_note_at = _latest_note_datetime(workorder)
    latest_note_text = _latest_note(workorder)
    recommendations = _recommendations_summary(workorder)
    close_date = workorder.updated_at if workorder.status == 'closed' else workorder.completed_at
    labor_hours = _resolved_workorder_labor_hours(workorder)
    labor_amount = _resolved_workorder_labor_amount(workorder)
    parts_amount = _resolved_workorder_parts_amount(workorder)
    misc_amount = Decimal('0')
    if invoice:
        misc_amount = (invoice.shop_supplies_fee or Decimal('0')) + (invoice.environmental_fee or Decimal('0'))

    row = {
        'Date In': _export_datetime(workorder.created_at),
        'Work Order Number': workorder.work_order_number,
        'Registration Number': getattr(vehicle, 'license_plate', '') or '',
        'Customer Code': getattr(customer, 'customer_number', '') or '',
        'Customer Name': _customer_name(customer),
        'Customer Balance': _export_decimal(invoice.amount_due if invoice else getattr(customer, 'current_balance', '')),
        'Job Status': workorder.get_status_display(),
        'Phone Number': _customer_phone(customer),
        'Mileage In': workorder.odometer_in or '',
        'Job Description': workorder.customer_concerns or '',
        'Contact Name': _contact_name(customer),
        'Received By': _full_name(workorder.created_by),
        'Needs By': _export_datetime(workorder.estimated_completion),
        'Payment Method': _payment_method(customer),
        'Created By': _full_name(workorder.created_by),
        'Work Hours': _export_decimal(labor_hours),
        'Labor Amount': _export_decimal(labor_amount),
        'Parts Amount': _export_decimal(parts_amount),
        'Sublet Amount': _export_decimal(invoice.sublet_subtotal if invoice else ''),
        'Misc Amount': _export_decimal(misc_amount),
        'Paid Amount': _export_decimal(invoice.amount_paid if invoice else ''),
        'Tax Amount': _export_decimal(invoice.tax_amount if invoice else ''),
        'Date Out': _export_datetime(workorder.completed_at),
        'Last Note Date': _export_datetime(latest_note_at),
        'Last Note': latest_note_text,
        'Status Date': _export_datetime(workorder.updated_at),
        'VAT Rate': tax1_rate,
        'Other Tax Rate': tax2_rate,
        'User Code': getattr(workorder.created_by, 'employee_id', '') or getattr(workorder.created_by, 'username', '') or '',
        'Changed Date': _export_datetime(workorder.updated_at),
        'Recommendations': recommendations,
        'Regular Service': _export_bool(workorder.maintenance_type == 'routine'),
        'Region Code': getattr(branch, 'code', '') or '',
        'Region': getattr(branch, 'name', '') or '',
        'Follow Up': _export_bool(bool(getattr(vehicle, 'next_service_due_date', None) or getattr(vehicle, 'next_service_due_mileage', None))),
        'Feedback': workorder.customer_feedback or '',
        'Account To Bill': _export_bool(getattr(customer, 'customer_type', '') in {'business', 'fleet'}) if customer else '',
        'Service Category': getattr(getattr(workorder, 'service_type', None), 'name', '') or workorder.get_maintenance_type_display(),
        'Invoice Status': invoice.get_status_display() if invoice else '',
        'Job Type': workorder.get_maintenance_type_display(),
        'Approval/Diagnosis Comment': workorder.approval_notes or workorder.diagnosis_notes or '',
        'Repair By': _full_name(workorder.lead_technician),
        'Job Location': getattr(branch, 'full_address', '') or getattr(branch, 'name', '') or '',
        'Insurance Provider': getattr(customer, 'insurance_provider', '') or '',
        'Next Action By': _full_name(workorder.service_coordinator),
        'Next Action On': _export_date(getattr(vehicle, 'next_service_due_date', None)),
        'Job Notes': workorder.special_instructions or '',
        'Parts Required': _parts_required(workorder),
        'Parts Progress': _parts_progress(workorder),
        'Close Date': _export_datetime(close_date),
        'Approval Requested Date': _export_datetime(workorder.approval_requested_at),
        'Approved Date': _export_datetime(workorder.approved_at),
        'Job Hours': _export_decimal(labor_hours),
        'Invoice Comment': (invoice.customer_notes or invoice.notes) if invoice else '',
        'Customer Score': workorder.customer_rating or '',
        'Approval Status': _approval_summary(workorder),
        'Subcontract Amount': _export_decimal(invoice.sublet_subtotal if invoice else ''),
        'Test Drive By': _full_name(getattr(triage, 'performed_by', None)),
        'Quality/Diagnosis Comment': workorder.quality_check_notes or workorder.diagnosis_notes or '',
        'Return Reference Number': getattr(workorder.related_work_order, 'work_order_number', '') if workorder.related_work_order else '',
        'Invoice Date': _export_date(getattr(invoice, 'invoice_date', None)),
        'Mileage In/Out': f"{workorder.odometer_in or ''}/{workorder.odometer_out or ''}",
        'Make': getattr(vehicle, 'make', '') or '',
        'Model': getattr(vehicle, 'model', '') or '',
    }

    return [row.get(header, '') for header in WORKORDER_EXPORT_HEADERS]


def _get_workorder_or_404(request, queryset=None, **lookup):
    """Retrieve a work order scoped to the current user's branch access."""
    base_queryset = queryset if queryset is not None else WorkOrder.objects.all()
    scoped_queryset = filter_queryset_for_user_branches(base_queryset, request.user, request=request, use_active_branch=True)
    return get_object_or_404(scoped_queryset, **lookup)


@login_required
def workorder_list_view(request):
    """
    Work order list with filtering and search
    """
    workorders = WorkOrder.objects.select_related(
        'customer', 'vehicle', 'primary_technician', 'branch'
    ).prefetch_related('assigned_technicians', 'tasks', 'parts').annotate(
        completed_tasks_count=Count('tasks', filter=Q(tasks__status='completed')),
        total_parts_quantity=Sum('parts__quantity'),
        total_hours=Sum('time_logs__duration_hours'),
        total_labor_cost=Sum('time_logs__labor_cost')
    )

    # Filter by active branch from session
    workorders = filter_queryset_for_user_branches(workorders, request.user, request=request, use_active_branch=True)
    
    # Apply filters
    status_filter = request.GET.get('status')
    priority_filter = request.GET.get('priority')
    technician_filter = request.GET.get('technician')
    search_query = request.GET.get('search')
    
    if status_filter:
        workorders = workorders.filter(status=status_filter)
    
    if priority_filter:
        workorders = workorders.filter(priority=priority_filter)
    
    if technician_filter:
        workorders = workorders.filter(primary_technician__id=technician_filter)
    
    if search_query:
        workorders = workorders.filter(
            Q(work_order_number__icontains=search_query) |
            Q(customer__user__first_name__icontains=search_query) |
            Q(customer__user__last_name__icontains=search_query) |
            Q(vehicle__make__icontains=search_query) |
            Q(vehicle__model__icontains=search_query) |
            Q(vehicle__license_plate__icontains=search_query) |
            Q(customer_concerns__icontains=search_query)
        )
    
    # Add annotations for counts
    workorders = workorders.annotate(
        task_count=Count('tasks'),
        parts_count=Count('parts'),
        total_labor_hours=Sum('tasks__actual_hours')
    )
    
    # Pagination
    paginator = Paginator(workorders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    technicians = User.objects.filter(role__in=['technician', 'manager', 'service_coordinator']).order_by('first_name')
    
    context = {
        'page_obj': page_obj,
        'workorders': page_obj,
        'status_choices': WorkOrder.STATUS_CHOICES,
        'priority_choices': WorkOrder.PRIORITY_CHOICES,
        'technicians': technicians,
        'current_filters': {
            'status': status_filter,
            'priority': priority_filter,
            'technician': technician_filter,
            'search': search_query,
        }
    }
    
    return render(request, 'workorders/workorder_list.html', context)


@login_required
def workorder_kanban_view(request):
    """
    Kanban board view of work orders
    """
    # Get work orders grouped by status
    workorders = WorkOrder.objects.select_related(
        'customer', 'vehicle', 'primary_technician', 'branch'
    ).prefetch_related('assigned_technicians', 'tasks', 'parts').annotate(
        completed_tasks_count=Count('tasks', filter=Q(tasks__status='completed')),
        task_count=Count('tasks'),
        total_parts_quantity=Sum('parts__quantity'),
        total_hours=Sum('time_logs__duration_hours'),
        total_labor_cost=Sum('time_logs__labor_cost')
    )

    # Filter by active branch from session
    workorders = filter_queryset_for_user_branches(workorders, request.user, request=request, use_active_branch=True)
    
    # Apply filters
    technician_filter = request.GET.get('technician')
    priority_filter = request.GET.get('priority')
    
    if technician_filter:
        workorders = workorders.filter(primary_technician__id=technician_filter)
    
    if priority_filter:
        workorders = workorders.filter(priority=priority_filter)
    
    # Group by status
    status_groups = {}
    for status_code, status_label in WorkOrder.STATUS_CHOICES:
        status_workorders = workorders.filter(status=status_code).annotate(
            task_count=Count('tasks'),
            completed_tasks=Count('tasks', filter=Q(tasks__status='completed'))
        )
        status_groups[status_code] = {
            'label': status_label,
            'workorders': status_workorders,
            'count': status_workorders.count()
        }
    
    # Get filter options
    technicians = User.objects.filter(role__in=['technician', 'manager', 'service_coordinator']).order_by('first_name')
    
    context = {
        'status_groups': status_groups,
        'priority_choices': WorkOrder.PRIORITY_CHOICES,
        'technicians': technicians,
        'current_filters': {
            'technician': technician_filter,
            'priority': priority_filter,
        }
    }
    
    return render(request, 'workorders/workorder_kanban.html', context)


@login_required
def workorder_detail_view(request, pk):
    """
    Work order detail view - redirects to Next.js frontend
    """
    # Verify work order exists and user has access
    queryset = WorkOrder.objects.select_related(
        'customer', 'vehicle', 'appointment', 'primary_technician', 'created_by', 'branch'
    )
    queryset = filter_queryset_for_user_branches(queryset, request.user, request=request, use_active_branch=True)
    workorder = _get_workorder_or_404(request, queryset=queryset, pk=pk)
    
    # Redirect to Next.js frontend
    from django.conf import settings
    frontend_url = getattr(settings, 'FRONTEND_BASE_URL', 'http://localhost:3000')
    return redirect(f'{frontend_url}/workorders/{pk}')


@login_required
def workorder_create_view(request):
    """
    Create new work order
    """
    customers = Customer.objects.select_related('user').order_by('user__first_name')
    vehicles = Vehicle.objects.select_related('owner').order_by('make', 'model')
    technicians = User.objects.filter(role__in=['technician', 'manager', 'service_coordinator']).order_by('first_name')
    
    if request.method == 'POST':
        branch = resolve_branch(request, branch_id=request.POST.get('branch') or request.POST.get('branch_id'))

        if branch is None:
            messages.error(request, 'Unable to determine branch for this work order.')
            return redirect('workorders:list')

        try:
            # Create work order
            workorder = WorkOrder.objects.create(
                customer_id=request.POST.get('customer'),
                vehicle_id=request.POST.get('vehicle'),
                primary_technician_id=request.POST.get('primary_technician'),
                priority=request.POST.get('priority', 'normal'),
                customer_concerns=request.POST.get('customer_concerns', ''),
                special_instructions=request.POST.get('special_instructions', ''),
                odometer_in=request.POST.get('odometer_in') or 0,
                estimated_completion=request.POST.get('estimated_completion') or None,
                created_by=request.user,
                branch=branch
            )
            
            # Add assigned technicians
            assigned_technician_ids = request.POST.getlist('assigned_technicians')
            if assigned_technician_ids:
                workorder.assigned_technicians.set(assigned_technician_ids)
            
            messages.success(request, f'Work order {workorder.work_order_number} created successfully!')
            return redirect('workorders:detail', pk=workorder.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating work order: {str(e)}')
    
    context = {
        'customers': customers,
        'vehicles': vehicles,
        'technicians': technicians,
        'priority_choices': WorkOrder.PRIORITY_CHOICES,
    }
    
    return render(request, 'workorders/workorder_create.html', context)


@login_required
def workorder_edit_view(request, pk):
    """
    Edit work order
    """
    workorder = get_object_or_404(
        filter_queryset_for_user_branches(WorkOrder.objects.all(), request.user, request=request, use_active_branch=True),
        pk=pk
    )
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'receptionist']:
        messages.error(request, 'You do not have permission to edit work orders.')
        return redirect('workorders:detail', pk=workorder.pk)
    
    customers = Customer.objects.select_related('user').order_by('user__first_name')
    vehicles = Vehicle.objects.select_related('owner').order_by('make', 'model')
    technicians = User.objects.filter(role__in=['technician', 'manager', 'service_coordinator']).order_by('first_name')
    
    if request.method == 'POST':
        try:
            # Update work order
            workorder.customer_id = request.POST.get('customer')
            workorder.vehicle_id = request.POST.get('vehicle')
            workorder.primary_technician_id = request.POST.get('primary_technician')
            workorder.priority = request.POST.get('priority', 'normal')
            workorder.status = request.POST.get('status', workorder.status)
            workorder.customer_concerns = request.POST.get('customer_concerns', '')
            workorder.special_instructions = request.POST.get('special_instructions', '')
            workorder.diagnosis_notes = request.POST.get('diagnosis_notes', '')
            odometer_in = request.POST.get('odometer_in')
            if odometer_in:
                workorder.odometer_in = int(odometer_in)
            workorder.estimated_completion = request.POST.get('estimated_completion') or None
            
            workorder.save()
            
            # Update assigned technicians
            assigned_technician_ids = request.POST.getlist('assigned_technicians')
            workorder.assigned_technicians.set(assigned_technician_ids)
            
            messages.success(request, f'Work order {workorder.work_order_number} updated successfully!')
            return redirect('workorders:detail', pk=workorder.pk)
            
        except Exception as e:
            messages.error(request, f'Error updating work order: {str(e)}')
    
    context = {
        'workorder': workorder,
        'customers': customers,
        'vehicles': vehicles,
        'technicians': technicians,
        'priority_choices': WorkOrder.PRIORITY_CHOICES,
        'status_choices': WorkOrder.STATUS_CHOICES,
    }
    
    return render(request, 'workorders/workorder_edit.html', context)


@login_required
def workorder_print_view(request, pk):
    """
    Printable work order view
    """
    queryset = WorkOrder.objects.select_related(
        'customer', 'vehicle', 'primary_technician'
    ).prefetch_related(
        'assigned_technicians', 'tasks', 'parts', 'time_logs'
    )

    workorder = _get_workorder_or_404(request, queryset=queryset, pk=pk)
    
    # Calculate totals
    total_labor_cost = workorder.tasks.aggregate(Sum('labor_cost'))['labor_cost__sum'] or 0
    total_parts_cost = workorder.parts.aggregate(Sum('selling_price'))['selling_price__sum'] or 0
    total_hours = workorder.tasks.aggregate(Sum('actual_hours'))['actual_hours__sum'] or 0
    
    context = {
        'workorder': workorder,
        'total_labor_cost': total_labor_cost,
        'total_parts_cost': total_parts_cost,
        'total_hours': total_hours,
        'print_generated_at': timezone.now(),
        'print_branch': workorder.branch,
    }
    
    return render(request, 'workorders/workorder_print.html', context)





# AJAX Views for dynamic updates

@login_required
def update_workorder_status(request, pk):
    """
    AJAX endpoint to update work order status
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    from django.core.exceptions import ValidationError
    
    workorder = _get_workorder_or_404(request, pk=pk)
    new_status = request.POST.get('status')
    
    if new_status not in dict(WorkOrder.STATUS_CHOICES):
        return JsonResponse({'error': 'Invalid status'}, status=400)
    
    try:
        workorder.transition_to(new_status, user=request.user)
        return JsonResponse({
            'success': True,
            'status': new_status,
            'status_display': workorder.get_status_display()
        })
    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def get_customer_vehicles(request, customer_id):
    """
    AJAX endpoint to get vehicles for a customer
    """
    vehicles = Vehicle.objects.filter(owner_id=customer_id).values(
        'id', 'make', 'model', 'year', 'license_plate', 'vin'
    )
    
    vehicle_list = []
    for vehicle in vehicles:
        vehicle_list.append({
            'id': vehicle['id'],
            'display': f"{vehicle['year']} {vehicle['make']} {vehicle['model']} - {vehicle['license_plate']}",
            'vin': vehicle['vin']
        })
    
    return JsonResponse({'vehicles': vehicle_list})


@login_required
def add_workorder_note(request, pk):
    """
    AJAX endpoint to add a note to work order
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    workorder = _get_workorder_or_404(request, pk=pk)
    
    try:
        data = json.loads(request.body)
        note_type = data.get('note_type', 'general')
        note_text = data.get('note', '').strip()
        is_important = data.get('is_important', False)
        is_customer_visible = data.get('is_customer_visible', False)
        
        if not note_text:
            return JsonResponse({'error': 'Note text is required'}, status=400)
        
        note = WorkOrderNote.objects.create(
            work_order=workorder,
            note_type=note_type,
            note=note_text,
            is_important=is_important,
            is_customer_visible=is_customer_visible,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'note': {
                'id': note.id,
                'note_type': note.get_note_type_display(),
                'note': note.note,
                'is_important': note.is_important,
                'is_customer_visible': note.is_customer_visible,
                'created_by': f"{note.created_by.first_name} {note.created_by.last_name}",
                'created_at': note.created_at.strftime('%Y-%m-%d %H:%M')
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def technician_time_clock(request, pk):
    """
    AJAX endpoint for technician time clock in/out
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    workorder = _get_workorder_or_404(request, pk=pk)
    
    try:
        data = json.loads(request.body)
        action = data.get('action')  # 'clock_in' or 'clock_out'
        task_id = data.get('task_id')  # optional
        description = data.get('description', '')
        
        if action == 'clock_in':
            # Check if already clocked in
            active_log = TechnicianTimeLog.objects.filter(
                work_order=workorder,
                technician=request.user,
                clock_out__isnull=True
            ).first()
            
            if active_log:
                return JsonResponse({'error': 'Already clocked in'}, status=400)
            
            time_log = TechnicianTimeLog.objects.create(
                work_order=workorder,
                task_id=task_id if task_id else None,
                technician=request.user,
                clock_in=timezone.now(),
                description=description,
                hourly_rate=request.user.hourly_rate if hasattr(request.user, 'hourly_rate') else 0
            )
            
            return JsonResponse({
                'success': True,
                'action': 'clocked_in',
                'time_log_id': time_log.id,
                'clock_in': time_log.clock_in.strftime('%Y-%m-%d %H:%M:%S')
            })
            
        elif action == 'clock_out':
            time_log_id = data.get('time_log_id')
            if time_log_id:
                time_log = get_object_or_404(TechnicianTimeLog, id=time_log_id, technician=request.user)
            else:
                # Find active time log
                time_log = TechnicianTimeLog.objects.filter(
                    work_order=workorder,
                    technician=request.user,
                    clock_out__isnull=True
                ).first()
                
            if not time_log:
                return JsonResponse({'error': 'No active time log found'}, status=400)
            
            time_log.clock_out = timezone.now()
            time_log.notes = data.get('notes', '')
            time_log.save()  # This will trigger duration and cost calculation
            
            return JsonResponse({
                'success': True,
                'action': 'clocked_out',
                'duration_hours': float(time_log.duration_hours),
                'labor_cost': float(time_log.labor_cost)
            })
            
        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def workorder_export_view(request):
    """
    Export work orders to Excel or CSV format
    """
    format_type = request.GET.get('format', 'xlsx').lower()
    
    workorders = WorkOrder.objects.select_related(
        'customer__user',
        'vehicle',
        'primary_technician',
        'branch',
        'created_by',
        'service_coordinator',
        'service_type',
        'triage_form__performed_by',
        'diagnosis',
    ).prefetch_related(
        'notes',
        'parts',
        'invoices',
        'diagnosis__repair_recommendations',
    ).all()
    
    # Filter by active branch from session
    workorders = filter_queryset_for_user_branches(workorders, request.user, request=request, use_active_branch=True)
    
    # Apply filters if provided
    status = request.GET.get('status')
    priority = request.GET.get('priority')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if status:
        workorders = workorders.filter(status=status)
    
    if priority:
        workorders = workorders.filter(priority=priority)
    
    if date_from:
        workorders = workorders.filter(created_at__date__gte=date_from)
    
    if date_to:
        workorders = workorders.filter(created_at__date__lte=date_to)
    
    if format_type == 'xlsx':
        return export_workorders_excel(workorders)
    if format_type == 'csv':
        return export_workorders_csv(workorders)
    else:
        return JsonResponse({'error': 'Invalid format. Use xlsx or csv.'}, status=400)


def export_workorders_excel(workorders):
    """Export work orders to Excel format."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Work Orders'

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for index, header in enumerate(WORKORDER_EXPORT_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    tax1_rate, tax2_rate = _tax_rates_for_export()
    for row_index, workorder in enumerate(workorders, start=2):
        row_values = _workorder_export_row(workorder, tax1_rate=tax1_rate, tax2_rate=tax2_rate)
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    worksheet.freeze_panes = 'A2'
    for column_cells in worksheet.columns:
        values = [str(cell.value or '') for cell in column_cells[:50]]
        max_length = max((len(value) for value in values), default=0)
        adjusted_width = min(max(max_length + 2, 12), 40)
        worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="workorders_export.xlsx"'
    return response


def export_workorders_csv(workorders):
    """Export work orders to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="workorders_export.csv"'
    
    writer = csv.writer(response)

    writer.writerow(WORKORDER_EXPORT_HEADERS)

    tax1_rate, tax2_rate = _tax_rates_for_export()

    for wo in workorders:
        writer.writerow(_workorder_export_row(wo, tax1_rate=tax1_rate, tax2_rate=tax2_rate))
    
    return response


@login_required
@require_http_methods(["POST"])
def add_task(request, pk):
    """
    Add a new service task to a work order
    """
    workorder = _get_workorder_or_404(request, pk=pk)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'technician']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        task_type = request.POST.get('task_type', 'repair')
        description = request.POST.get('description', '').strip()
        detailed_notes = request.POST.get('detailed_notes', '').strip()
        assigned_to_id = request.POST.get('assigned_to')
        estimated_hours = request.POST.get('estimated_hours', 0)
        labor_rate = request.POST.get('labor_rate', 0)
        
        if not description:
            return JsonResponse({'success': False, 'error': 'Description is required'})
        
        # Get next sequence order
        max_sequence = workorder.tasks.aggregate(max_seq=Max('sequence_order'))['max_seq']
        next_sequence = (max_sequence or 0) + 1
        
        # Create task
        task = ServiceTask.objects.create(
            work_order=workorder,
            task_type=task_type,
            description=description,
            detailed_notes=detailed_notes,
            assigned_to_id=assigned_to_id if assigned_to_id else None,
            estimated_hours=float(estimated_hours) if estimated_hours else 0,
            labor_rate=float(labor_rate) if labor_rate else 0,
            sequence_order=next_sequence
        )
        
        return JsonResponse({
            'success': True, 
            'message': f'Task "{task.description}" added successfully',
            'task_id': task.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def update_task_status(request, pk, task_id):
    """
    Update the status of a service task
    """
    workorder = _get_workorder_or_404(request, pk=pk)
    task = get_object_or_404(ServiceTask, id=task_id, work_order=workorder)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'technician']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        new_status = request.POST.get('status')
        if new_status not in dict(ServiceTask.STATUS_CHOICES):
            return JsonResponse({'success': False, 'error': 'Invalid status'})
        
        # Update timestamps based on status
        if new_status == 'in_progress' and task.status != 'in_progress':
            task.started_at = timezone.now()
        elif new_status == 'completed' and task.status != 'completed':
            task.completed_at = timezone.now()
            # If no actual hours recorded, use estimated hours
            if task.actual_hours == 0 and task.estimated_hours > 0:
                task.actual_hours = task.estimated_hours
        
        task.status = new_status
        task.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Task status updated to {task.get_status_display()}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["DELETE", "POST"])
def delete_task(request, pk, task_id):
    """
    Delete a service task
    """
    workorder = _get_workorder_or_404(request, pk=pk)
    task = get_object_or_404(ServiceTask, id=task_id, work_order=workorder)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        task_description = task.description
        task.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Task "{task_description}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_task(request, pk, task_id):
    """Get task details for editing"""
    try:
        workorder = _get_workorder_or_404(request, pk=pk)
        task = get_object_or_404(ServiceTask, id=task_id, work_order=workorder)
        
        return JsonResponse({
            'success': True,
            'task': {
                'id': task.id,
                'task_type': task.task_type,
                'description': task.description,
                'detailed_notes': task.detailed_notes or '',
                'assigned_to': task.assigned_to.id if task.assigned_to else None,
                'estimated_hours': float(task.estimated_hours) if task.estimated_hours else 0,
                'labor_rate': float(task.labor_rate) if task.labor_rate else 0,
                'status': task.status,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def update_task(request, pk, task_id):
    """Update an existing service task"""
    try:
        workorder = _get_workorder_or_404(request, pk=pk)
        task = get_object_or_404(ServiceTask, id=task_id, work_order=workorder)
        
        # Check permissions
        if request.user.role not in ['admin', 'manager', 'technician']:
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        
        # Update task fields
        task.task_type = request.POST.get('task_type')
        task.description = request.POST.get('description')
        task.detailed_notes = request.POST.get('detailed_notes', '')
        
        assigned_to_id = request.POST.get('assigned_to')
        if assigned_to_id:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            task.assigned_to = User.objects.get(id=assigned_to_id)
        else:
            task.assigned_to = None
        
        estimated_hours = request.POST.get('estimated_hours')
        if estimated_hours:
            task.estimated_hours = Decimal(estimated_hours)
        
        labor_rate = request.POST.get('labor_rate')
        if labor_rate:
            task.labor_rate = Decimal(labor_rate)
        
        # Calculate labor cost
        if task.estimated_hours and task.labor_rate:
            task.labor_cost = task.estimated_hours * task.labor_rate
        
        task.save()
        
        # Create activity note
        WorkOrderNote.objects.create(
            work_order=workorder,
            note_type='task',
            note=f'Task updated: {task.description}',
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Task updated successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def update_diagnosis(request, pk):
    """
    Update diagnosis notes and estimate for a work order
    """
    workorder = _get_workorder_or_404(request, pk=pk)
    
    # Check if user can perform diagnosis
    if request.user.role not in ['admin', 'manager', 'technician']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        # Update diagnosis information
        workorder.diagnosis_notes = request.POST.get('diagnosis_notes', '').strip()
        workorder.diagnosis_by = request.user
        workorder.diagnosis_completed_at = timezone.now()
        
        # Update labor estimates
        estimated_labor_hours = request.POST.get('estimated_labor_hours')
        if estimated_labor_hours:
            workorder.estimated_labor_hours = Decimal(estimated_labor_hours)
        
        estimated_labor_cost = request.POST.get('estimated_labor_cost')
        if estimated_labor_cost:
            workorder.estimated_labor_cost = Decimal(estimated_labor_cost)
        
        workorder.save()
        
        # Create activity note
        WorkOrderNote.objects.create(
            work_order=workorder,
            note_type='diagnosis',
            note=f'Diagnosis completed: {workorder.diagnosis_notes[:100]}...',
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Diagnosis updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def add_part(request, pk):
    """
    Add a part to the work order estimate
    """
    workorder = _get_workorder_or_404(request, pk=pk)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'technician']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        # Get form data
        part_number = request.POST.get('part_number', '').strip()
        part_name = request.POST.get('part_name', '').strip()
        description = request.POST.get('description', '').strip()
        quantity = Decimal(request.POST.get('quantity', '1'))
        unit_cost = Decimal(request.POST.get('unit_cost', '0'))
        markup_percentage = Decimal(request.POST.get('markup_percentage', '0'))
        task_id = request.POST.get('task_id', '').strip()
        
        # Validate required fields
        if not part_number or not part_name:
            return JsonResponse({'success': False, 'error': 'Part number and name are required'})
        
        # Get task if provided
        task = None
        if task_id:
            try:
                task = ServiceTask.objects.get(id=task_id, work_order=workorder)
            except ServiceTask.DoesNotExist:
                pass
        
        # Calculate costs
        total_cost = quantity * unit_cost
        selling_price = total_cost * (Decimal('1') + markup_percentage / Decimal('100'))
        
        # Create the part
        part = WorkOrderPart.objects.create(
            work_order=workorder,
            task=task,
            part_number=part_number,
            part_name=part_name,
            description=description,
            quantity=quantity,
            unit_cost=unit_cost,
            total_cost=total_cost,
            markup_percentage=markup_percentage,
            selling_price=selling_price,
            status='pending'
        )
        
        # Update work order estimated parts cost
        total = workorder.parts.aggregate(total=Sum('selling_price'))['total']
        workorder.estimated_parts_cost = Decimal(str(total)) if total is not None else Decimal('0')
        workorder.save()
        
        # Create activity note
        WorkOrderNote.objects.create(
            work_order=workorder,
            note_type='parts',
            note=f'Added part: {part_name} (Qty: {quantity})',
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Part "{part_name}" added successfully',
            'part_id': part.id
        })
        
    except ValueError as e:
        return JsonResponse({'success': False, 'error': 'Invalid number format'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["DELETE"])
def delete_part(request, pk, part_id):
    """
    Delete a part from the work order
    """
    workorder = _get_workorder_or_404(request, pk=pk)
    part = get_object_or_404(WorkOrderPart, id=part_id, work_order=workorder)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'technician']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        part_name = part.part_name
        part.delete()
        
        # Update work order estimated parts cost
        total = workorder.parts.aggregate(total=Sum('selling_price'))['total']
        workorder.estimated_parts_cost = Decimal(str(total)) if total is not None else Decimal('0')
        workorder.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Part "{part_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def update_part_status(request, pk, part_id):
    """
    Update the status of a part (pending, ordered, received, installed)
    """
    workorder = _get_workorder_or_404(request, pk=pk)
    part = get_object_or_404(WorkOrderPart, id=part_id, work_order=workorder)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'technician']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if new_status not in ['pending', 'ordered', 'received', 'installed', 'returned']:
            return JsonResponse({'success': False, 'error': 'Invalid status'})
        
        # Update status and tracking fields
        part.status = new_status
        
        if new_status == 'ordered' and not part.ordered_at:
            part.ordered_at = timezone.now()
        elif new_status == 'received' and not part.received_at:
            part.received_at = timezone.now()
        elif new_status == 'installed':
            part.installed_at = timezone.now()
            part.installed_by = request.user
        
        part.save()
        
        # Create activity note
        WorkOrderNote.objects.create(
            work_order=workorder,
            note_type='parts',
            note=f'Part "{part.part_name}" status changed to {part.get_status_display()}',
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Part status updated to {part.get_status_display()}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def request_approval(request, pk):
    """
    Submit estimate for customer approval
    """
    workorder = _get_workorder_or_404(request, pk=pk)
    
    # Check permissions
    if request.user.role not in ['admin', 'manager', 'technician']:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        # Validate that diagnosis is complete
        if not workorder.diagnosis_notes:
            return JsonResponse({
                'success': False, 
                'error': 'Please complete the diagnosis notes before requesting approval'
            })

        if workorder.status != 'diagnosis':
            return JsonResponse({
                'success': False,
                'error': f'Work order must be in diagnosis before requesting approval. Current status: {workorder.get_status_display()}'
            })
        
        # Check if estimate already exists
        if hasattr(workorder, 'estimate') and workorder.estimate:
            # Update existing estimate status
            estimate = workorder.estimate
            estimate.status = 'sent'
            estimate.save()
        else:
            # Create an Estimate record from the work order
            from datetime import timedelta
            
            estimate = Estimate.objects.create(
                customer=workorder.customer,
                vehicle=workorder.vehicle,
                work_order=workorder,
                status='sent',
                estimate_date=timezone.now().date(),
                valid_until=(timezone.now() + timedelta(days=30)).date(),
                title=f"Estimate for {workorder.vehicle.year} {workorder.vehicle.make} {workorder.vehicle.model}",
                description=workorder.diagnosis_notes or workorder.customer_concerns,
                labor_subtotal=workorder.estimated_labor_cost or Decimal('0'),
                parts_subtotal=workorder.estimated_parts_cost or Decimal('0'),
                subtotal=workorder.estimated_total or Decimal('0'),
                total=workorder.estimated_total or Decimal('0'),
                created_by=request.user,
                sent_by=request.user,
            )
            
            # Create line items for labor (tasks)
            for task in workorder.tasks.all():
                EstimateLineItem.objects.create(
                    estimate=estimate,
                    item_type='labor',
                    description=f"{task.get_task_type_display()} - {task.description}",
                    quantity=task.estimated_hours or Decimal('1'),
                    unit_price=task.labor_rate or Decimal('0'),
                    total=(task.estimated_hours or Decimal('1')) * (task.labor_rate or Decimal('0')),
                )
            
            # Create line items for parts
            for part in workorder.parts.all():
                EstimateLineItem.objects.create(
                    estimate=estimate,
                    item_type='part',
                    description=f"{part.part_name} ({part.part_number})",
                    quantity=part.quantity,
                    unit_price=part.unit_cost,
                    total=part.selling_price,
                )
        
        workorder.requires_approval = True
        workorder.approval_requested_at = timezone.now()
        workorder.save(update_fields=['requires_approval', 'approval_requested_at'])
        workorder.transition_to('awaiting_approval', user=request.user)
        
        # Create activity note
        WorkOrderNote.objects.create(
            work_order=workorder,
            note_type='status',
            note=f'Estimate #{estimate.estimate_number} submitted for customer approval. Total: ${workorder.estimated_total}',
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Estimate #{estimate.estimate_number} submitted for customer approval'
        })
        
    except Exception as e:
        import logging
        import traceback
        logging.getLogger(__name__).error(
            "Error submitting estimate for approval: %s\n%s", e, traceback.format_exc(), exc_info=True
        )
        from django.conf import settings
        msg = str(e) if settings.DEBUG else 'An error occurred while submitting the estimate.'
        return JsonResponse({'success': False, 'error': msg})


@login_required
def workorder_get_part(request, pk, part_id):
    """Get part details for editing"""
    try:
        workorder = _get_workorder_or_404(request, pk=pk)
        part = get_object_or_404(WorkOrderPart, pk=part_id, work_order=workorder)
        
        return JsonResponse({
            'success': True,
            'part': {
                'id': part.id,
                'part_number': part.part_number,
                'part_name': part.part_name,
                'description': part.description or '',
                'quantity': part.quantity,
                'unit_cost': float(part.unit_cost),
                'markup_percentage': part.markup_percentage or 30,
                'total_cost': float(part.total_cost),
                'status': part.status,
                'task_id': part.task.id if part.task else None,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def workorder_update_part(request, pk, part_id):
    """Update an existing part in the work order"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        workorder = _get_workorder_or_404(request, pk=pk)
        part = get_object_or_404(WorkOrderPart, pk=part_id, work_order=workorder)
        
        # Update part fields
        part.part_number = request.POST.get('part_number')
        part.part_name = request.POST.get('part_name')
        part.description = request.POST.get('description', '')
        part.quantity = Decimal(request.POST.get('quantity', '1'))
        part.unit_cost = Decimal(request.POST.get('unit_cost', '0'))
        part.markup_percentage = Decimal(request.POST.get('markup_percentage', '30'))
        
        # Handle task assignment
        task_id = request.POST.get('task_id', '').strip()
        if task_id:
            try:
                part.task = ServiceTask.objects.get(id=task_id, work_order=workorder)
            except ServiceTask.DoesNotExist:
                part.task = None
        else:
            part.task = None
        
        # Calculate total cost and selling price
        part.total_cost = part.unit_cost * part.quantity
        part.selling_price = part.total_cost * (Decimal('1') + part.markup_percentage / Decimal('100'))
        part.save()
        
        # Update work order estimated parts cost
        total = workorder.parts.aggregate(total=Sum('selling_price'))['total']
        workorder.estimated_parts_cost = Decimal(str(total)) if total is not None else Decimal('0')
        workorder.save()
        
        # Create activity note
        WorkOrderNote.objects.create(
            work_order=workorder,
            note_type='part',
            note=f'Part updated: {part.part_name} (Qty: {part.quantity})',
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Part updated successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
