"""Orchestration helpers for import batches."""
from __future__ import annotations

import json
import uuid
from typing import Any

from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from django.utils import timezone

from apps.accounts.admin_views import log_audit
from apps.data_exchange.models import ImportBatch, ImportRowResult
from apps.data_exchange.registry import get_exporter, get_importer


class CommitCancelled(Exception):
    """Raised when an import commit is cancelled by the user."""


def check_import_cancelled(opts: dict | None) -> None:
    """Raise CommitCancelled if the batch was cancelled (re-reads DB)."""
    if not opts:
        return
    batch_id = opts.get('_batch_id')
    if not batch_id:
        return
    row = (
        ImportBatch.objects.filter(pk=batch_id)
        .values('options', 'status')
        .first()
    )
    if not row:
        raise CommitCancelled('Import batch was deleted')
    if (row['options'] or {}).get('cancel_requested'):
        raise CommitCancelled('Cancelled by user')
    if row['status'] not in {
        ImportBatch.STATUS_COMMITTING,
        ImportBatch.STATUS_PREVIEWING,
        ImportBatch.STATUS_PREVIEWED,
        ImportBatch.STATUS_UPLOADED,
    }:
        # Cancel API marks failed immediately; stop the worker.
        if row['status'] == ImportBatch.STATUS_FAILED:
            raise CommitCancelled('Cancelled by user')


def cancel_batch(batch: ImportBatch) -> ImportBatch:
    """Request cancel for a running preview/commit and unlock the batch."""
    if batch.status not in {
        ImportBatch.STATUS_COMMITTING,
        ImportBatch.STATUS_PREVIEWING,
    }:
        raise ValueError(f'Cannot cancel import in status "{batch.status}"')

    opts = dict(batch.options or {})
    opts['cancel_requested'] = True
    batch.options = opts
    batch.status = ImportBatch.STATUS_FAILED
    batch.error_message = 'Cancelled by user'
    batch.save(update_fields=['options', 'status', 'error_message'])
    return batch


def _parse_options(raw: Any) -> dict:
    if raw is None or raw == '':
        return {}
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str):
        return json.loads(raw)
    raise ValueError('options must be a JSON object')


def create_batch(
    *,
    module_key: str,
    upload: UploadedFile,
    user,
    options: Any = None,
) -> ImportBatch:
    importer = get_importer(module_key)
    filename = getattr(upload, 'name', 'upload.xlsx')
    if not any(filename.lower().endswith(ext) for ext in importer.supported_extensions):
        raise ValueError(
            f'Unsupported file type for {importer.label}. '
            f'Allowed: {", ".join(importer.supported_extensions)}'
        )

    opts = {**importer.default_options(), **_parse_options(options)}
    batch = ImportBatch.objects.create(
        uuid=uuid.uuid4(),
        module_key=module_key,
        status=ImportBatch.STATUS_UPLOADED,
        source_file=upload,
        original_filename=filename,
        options=opts,
        created_by=user,
    )
    return batch


def _preview_options(batch: ImportBatch) -> dict:
    """Preview stays fast: skip remote VIN decode unless explicitly requested."""
    opts = dict(batch.options or {})
    opts['decode_vin_for_missing_fields'] = bool(opts.get('decode_vin_on_preview', False))
    return opts


def _commit_options(batch: ImportBatch) -> dict:
    """Commit uses the batch options as uploaded (no forced NHTSA decode)."""
    return dict(batch.options or {})


def _celery_workers_available() -> bool:
    """Return True only when at least one Celery worker responds to ping."""
    try:
        from config.celery import app as celery_app

        replies = celery_app.control.ping(timeout=0.5) or []
        return bool(replies)
    except Exception:
        return False


def _dispatch_background(task, *args, **kwargs) -> None:
    """
    Run work off the ASGI request thread.

    Prefer Celery when a worker is actually online; otherwise use a daemon
    thread so local/dev imports do not sit forever in Redis with no consumer.
    """
    import threading

    from django.conf import settings

    def _thread_apply():
        task.apply(args=args, kwargs=kwargs)

    force_thread = bool(getattr(settings, 'DATA_EXCHANGE_FORCE_THREAD', False))
    if (
        force_thread
        or getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
        or getattr(settings, 'DEBUG', False)
        or not _celery_workers_available()
    ):
        threading.Thread(target=_thread_apply, daemon=True).start()
        return

    try:
        task.delay(*args, **kwargs)
    except Exception:
        threading.Thread(target=_thread_apply, daemon=True).start()


def queue_preview(batch: ImportBatch) -> ImportBatch:
    """Mark batch previewing and run validation in the background (Celery or thread)."""
    if batch.status == ImportBatch.STATUS_COMMITTING:
        return batch
    # Allow re-queue when a previous background job died mid-preview.
    if batch.status == ImportBatch.STATUS_PREVIEWING and (batch.preview_report or {}):
        return batch

    batch.status = ImportBatch.STATUS_PREVIEWING
    batch.error_message = ''
    batch.save(update_fields=['status', 'error_message'])

    from apps.data_exchange.tasks import run_import_preview_task
    _dispatch_background(run_import_preview_task, batch.id)
    return batch


def queue_commit(batch: ImportBatch, *, force: bool = False) -> ImportBatch:
    """Queue commit in the background so Daphne/ASGI is not blocked."""
    if batch.status == ImportBatch.STATUS_COMPLETED:
        raise ValueError('This import batch was already committed')
    if batch.status == ImportBatch.STATUS_ROLLED_BACK:
        raise ValueError('This import batch was rolled back and cannot be committed again')
    if batch.status == ImportBatch.STATUS_PREVIEWING:
        raise ValueError('Preview is still running. Wait for validation to finish before committing.')
    if batch.status not in {
        ImportBatch.STATUS_UPLOADED,
        ImportBatch.STATUS_PREVIEWED,
        ImportBatch.STATUS_FAILED,
        ImportBatch.STATUS_COMMITTING,  # allow resume if worker died
    }:
        raise ValueError(f'Cannot commit import in status "{batch.status}"')

    report = batch.preview_report or {}
    if batch.status not in {
        ImportBatch.STATUS_PREVIEWED,
        ImportBatch.STATUS_COMMITTING,
    } and not force:
        raise ValueError('Run preview first and wait until status is "previewed".')
    if not report.get('can_commit') and not force:
        raise ValueError('Preview found nothing to import. Review the validation report first.')

    batch.status = ImportBatch.STATUS_COMMITTING
    batch.error_message = ''
    batch.save(update_fields=['status', 'error_message'])

    from apps.data_exchange.tasks import run_import_commit_task
    _dispatch_background(run_import_commit_task, batch.id, force=force)
    return batch


def run_preview(batch: ImportBatch, *, for_background: bool = False) -> ImportBatch:
    importer = get_importer(batch.module_key)
    if not for_background and batch.status != ImportBatch.STATUS_PREVIEWING:
        batch.status = ImportBatch.STATUS_PREVIEWING
        batch.save(update_fields=['status'])

    opts = _preview_options(batch)
    with batch.source_file.open('rb') as handle:
        result = importer.preview(handle, opts)
    report = result.as_dict()
    batch.mark_previewed(report)
    log_audit(
        user=batch.created_by,
        action='import',
        model_name='ImportBatch',
        object_repr=f'Import Preview: {batch.original_filename}',
        changes={
            'batch_id': batch.id,
            'module_key': batch.module_key,
            'summary': report.get('summary'),
            'error_count': report.get('error_count'),
            'warning_count': report.get('warning_count'),
            'filename': batch.original_filename,
            'mode': 'preview',
            'async': for_background,
        },
    )
    return batch


def run_commit(batch: ImportBatch, *, force: bool = False, for_background: bool = False) -> ImportBatch:
    if batch.status == ImportBatch.STATUS_COMPLETED:
        raise ValueError('This import batch was already committed')
    if batch.status == ImportBatch.STATUS_ROLLED_BACK:
        raise ValueError('This import batch was rolled back and cannot be committed again')

    # Background worker already moved status to committing via queue_commit
    if for_background:
        allowed = {
            ImportBatch.STATUS_COMMITTING,
            ImportBatch.STATUS_PREVIEWED,
            ImportBatch.STATUS_UPLOADED,
            ImportBatch.STATUS_FAILED,
        }
    else:
        allowed = {
            ImportBatch.STATUS_UPLOADED,
            ImportBatch.STATUS_PREVIEWED,
            ImportBatch.STATUS_FAILED,
            ImportBatch.STATUS_COMMITTING,
        }
    if batch.status not in allowed:
        raise ValueError(f'Cannot commit import in status "{batch.status}"')

    if batch.status in {ImportBatch.STATUS_UPLOADED, ImportBatch.STATUS_FAILED} and not force:
        run_preview(batch, for_background=for_background)
        batch.refresh_from_db()

    report = batch.preview_report or {}
    if not report.get('can_commit') and not force:
        raise ValueError('Preview found nothing to import. Review the validation report first.')

    importer = get_importer(batch.module_key)
    if batch.status != ImportBatch.STATUS_COMMITTING:
        batch.status = ImportBatch.STATUS_COMMITTING
        batch.save(update_fields=['status'])

    def _progress(summary, created_refs):
        check_import_cancelled({'_batch_id': batch.pk})
        ImportBatch.objects.filter(pk=batch.pk).update(
            summary={
                **(summary or {}),
                'progress': {
                    'customers_created': summary.get('customers_created', 0),
                    'vehicles_created': summary.get('vehicles_created', 0),
                    'parts_created': summary.get('parts_created', 0),
                    'staff_created': summary.get('staff_created', 0),
                },
            },
            created_object_refs=created_refs or {},
        )

    try:
        commit_opts = _commit_options(batch)
        commit_opts['_progress_callback'] = _progress
        commit_opts['_batch_id'] = batch.pk
        with batch.source_file.open('rb') as handle:
            result = importer.commit(handle, commit_opts)

        batch.refresh_from_db()
        if (batch.options or {}).get('cancel_requested') or batch.status == ImportBatch.STATUS_FAILED:
            if result.created_refs:
                ImportBatch.objects.filter(pk=batch.pk).update(
                    created_object_refs=result.created_refs,
                    summary={**(result.summary or {}), 'cancelled': True},
                )
            batch.refresh_from_db()
            if batch.status != ImportBatch.STATUS_FAILED:
                batch.mark_failed('Cancelled by user', summary=result.summary)
            return batch

        # Persist row results (cap for very large imports; prefer problems over info spam)
        ImportRowResult.objects.filter(batch=batch).delete()
        prioritized = [
            i for i in result.issues
            if i.level in {'error', 'warning'}
        ] + [
            i for i in result.issues
            if i.level == 'info'
        ]
        rows_to_create = []
        for issue in prioritized[:5000]:
            rows_to_create.append(ImportRowResult(
                batch=batch,
                row_number=issue.row_number or 0,
                entity_type=issue.entity_type or ImportRowResult.ENTITY_OTHER,
                action=issue.action if issue.action in {
                    ImportRowResult.ACTION_CREATE,
                    ImportRowResult.ACTION_UPDATE,
                    ImportRowResult.ACTION_MATCH,
                    ImportRowResult.ACTION_SKIP,
                    ImportRowResult.ACTION_FAIL,
                } else ImportRowResult.ACTION_FAIL,
                identifier=issue.identifier or '',
                message=issue.message or '',
                object_id=(issue.payload or {}).get('customer_id')
                or (issue.payload or {}).get('vehicle_id'),
                payload=issue.payload or {},
            ))
        if rows_to_create:
            ImportRowResult.objects.bulk_create(rows_to_create, batch_size=500)

        summary = {
            **result.summary,
            'issue_count': len(result.issues),
        }
        batch.mark_completed(summary, result.created_refs)
        log_audit(
            user=batch.created_by,
            action='import',
            model_name='ImportBatch',
            object_repr=f'Excel Import: {batch.original_filename}',
            changes={
                'batch_id': batch.id,
                'module_key': batch.module_key,
                'imported': (
                    summary.get('vehicles_created', 0)
                    + summary.get('customers_created', 0)
                    + summary.get('rows_created', 0)
                ),
                'skipped': summary.get('vehicles_skipped', 0),
                'total_errors': summary.get('vehicles_failed', 0),
                'filename': batch.original_filename,
                'summary': summary,
                'mode': 'commit',
            },
        )
        return batch
    except CommitCancelled as exc:
        batch.refresh_from_db()
        if batch.status != ImportBatch.STATUS_FAILED:
            batch.mark_failed(str(exc))
        log_audit(
            user=batch.created_by,
            action='import',
            model_name='ImportBatch',
            object_repr=f'Excel Import Cancelled: {batch.original_filename}',
            changes={
                'batch_id': batch.id,
                'module_key': batch.module_key,
                'error': str(exc),
                'filename': batch.original_filename,
                'mode': 'commit_cancelled',
            },
        )
        return batch
    except Exception as exc:  # noqa: BLE001
        batch.mark_failed(str(exc))
        log_audit(
            user=batch.created_by,
            action='import',
            model_name='ImportBatch',
            object_repr=f'Excel Import Failed: {batch.original_filename}',
            changes={
                'batch_id': batch.id,
                'module_key': batch.module_key,
                'error': str(exc),
                'filename': batch.original_filename,
                'mode': 'commit',
            },
        )
        raise


def run_rollback(batch: ImportBatch) -> ImportBatch:
    if batch.status != ImportBatch.STATUS_COMPLETED:
        raise ValueError('Only completed imports can be rolled back')
    if not batch.created_object_refs:
        raise ValueError('No created object references are available for rollback')

    importer = get_importer(batch.module_key)
    with transaction.atomic():
        result = importer.rollback(batch.created_object_refs)
        summary = {**(batch.summary or {}), 'rollback': result}
        batch.mark_rolled_back(summary)

    log_audit(
        user=batch.created_by,
        action='import',
        model_name='ImportBatch',
        object_repr=f'Import Rollback: {batch.original_filename}',
        changes={
            'batch_id': batch.id,
            'module_key': batch.module_key,
            'filename': batch.original_filename,
            'rollback': result,
            'mode': 'rollback',
            'rolled_back_at': timezone.now().isoformat(),
        },
    )
    return batch


def run_export(module_key: str, options: dict | None = None):
    exporter = get_exporter(module_key)
    return exporter.export(options or {})


def build_validation_report_workbook(batch: ImportBatch):
    """Re-run preview and return a full Excel validation report (all issues)."""
    from io import BytesIO

    import openpyxl

    importer = get_importer(batch.module_key)
    with batch.source_file.open('rb') as handle:
        result = importer.preview(handle, batch.options)

    workbook = openpyxl.Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'
    summary_sheet.append(['Field', 'Value'])
    summary_sheet.append(['filename', batch.original_filename])
    summary_sheet.append(['module', batch.module_key])
    summary_sheet.append(['format', result.format_detected])
    for key, value in sorted(result.summary.items()):
        summary_sheet.append([key, value])
    summary_sheet.append(['error_count', sum(1 for i in result.issues if i.level == 'error')])
    summary_sheet.append(['warning_count', sum(1 for i in result.issues if i.level == 'warning')])

    breakdown_sheet = workbook.create_sheet('Breakdown')
    breakdown_sheet.append(['level', 'code', 'count', 'example_message'])
    from apps.data_exchange.importers.base import _issue_breakdown
    for row in _issue_breakdown(result.issues):
        breakdown_sheet.append([row['level'], row['code'], row['count'], row['message']])

    issues_sheet = workbook.create_sheet('All Issues')
    issues_sheet.append([
        'row_number', 'level', 'code', 'entity_type', 'action', 'identifier', 'message',
    ])
    for issue in result.issues:
        issues_sheet.append([
            issue.row_number,
            issue.level,
            issue.code,
            issue.entity_type,
            issue.action,
            issue.identifier,
            issue.message,
        ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f'import_validation_batch_{batch.id}.xlsx'
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return buffer, filename, content_type
